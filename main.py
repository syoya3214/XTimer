import sys
import time
import threading
import keyboard
import configparser
import ctypes
import os
import openpyxl
import gc
import re
from PyQt5.QtWidgets import (QApplication, QWidget, QLabel, QMenu, QVBoxLayout,
                              QHBoxLayout, QTableWidget, QTableWidgetItem,
                              QStyledItemDelegate, QDesktopWidget, QPushButton,
                              QColorDialog)
from PyQt5.QtGui import (QFont, QColor, QPainter, QPen, QPainterPath,
                         QBrush, QFontMetrics, QIcon, QPixmap, QMovie)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QUrl, QObject, QTimer, QSize, QPoint
from PyQt5.QtMultimedia import QMediaPlayer, QMediaContent

# ============================================================
# config.ini 安全書き込みヘルパー
# ============================================================
def backup_config(config_path):
    bak = config_path + '.bak'
    try:
        if os.path.isfile(config_path):
            import shutil
            shutil.copy2(config_path, bak)
    except Exception as e:
        print(f"[backup_config] {e}")


def restore_config(config_path):
    bak = config_path + '.bak'
    if not os.path.isfile(bak):
        return False
    try:
        import shutil
        shutil.copy2(bak, config_path)
        return True
    except Exception as e:
        print(f"[restore_config] {e}")
        return False


def safe_set_config(config_path, key, value, section='Settings'):
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
    except FileNotFoundError:
        lines = [f'[{section}]\n']

    in_section = False
    key_found  = False
    section_end = len(lines)

    for i, line in enumerate(lines):
        stripped = line.strip()
        if stripped.startswith('['):
            if stripped.lower() == f'[{section.lower()}]':
                in_section = True
            elif in_section:
                section_end = i
                break
            continue
        if in_section:
            m = re.match(r'^(\s*)(' + re.escape(key) + r')(\s*[=:]\s*)', line, re.IGNORECASE)
            if m:
                lines[i] = f'{m.group(1)}{key} = {value}\n'
                key_found = True
                break

    if not key_found:
        lines.insert(section_end, f'{key} = {value}\n')

    backup_config(config_path)
    with open(config_path, 'w', encoding='utf-8') as f:
        f.writelines(lines)

# ============================================================
# フォルダ定数 / パスヘルパー
# ============================================================
SOUND_DIR = 'sound'
IMG_DIR   = 'img'
XLSX_DIR  = 'xlsx'

def sound_path(filename):
    return os.path.abspath(os.path.join(SOUND_DIR, filename))

def img_path(filename):
    return os.path.join(IMG_DIR, filename)

def xlsx_path(filename):
    return os.path.join(XLSX_DIR, filename)

IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.webp'}

def is_image_label(label_str):
    if not label_str:
        return False
    label_str = label_str.strip()
    ext = os.path.splitext(label_str)[1].lower()
    if ext not in IMAGE_EXTENSIONS:
        return False
    return os.path.isfile(img_path(label_str))

def _make_color_icon(color_str, size=16):
    pm = QPixmap(size, size)
    pm.fill(QColor(color_str) if color_str else QColor(0, 0, 0, 0))
    return QIcon(pm)

def _cursor_screen_center():
    desktop = QApplication.desktop()
    sg = desktop.screenGeometry(desktop.screenNumber(desktop.cursor().pos()))
    return sg, sg.center()


# ============================================================
# hotkey キャプチャスレッド
# ============================================================
class HotkeyCaptureThread(QThread):
    captured = pyqtSignal(str)

    _MOD_NORMALIZE = {
        'left ctrl': 'ctrl',   'right ctrl': 'ctrl',
        'left shift': 'shift', 'right shift': 'shift',
        'left alt': 'alt',     'right alt': 'alt',
        'left windows': 'windows', 'right windows': 'windows',
    }
    _MODIFIERS = {'ctrl', 'shift', 'alt', 'windows'}

    @classmethod
    def _norm(cls, name):
        return cls._MOD_NORMALIZE.get(name, name)

    def run(self):
        import threading as _threading
        pressed  = []
        done     = _threading.Event()
        hook_ref = [None]

        def on_event(e):
            if done.is_set(): return
            raw  = (e.name or '').lower()
            name = self._norm(raw)
            if not name: return
            if e.event_type == keyboard.KEY_DOWN:
                if name not in pressed:
                    pressed.append(name)
            elif e.event_type == keyboard.KEY_UP:
                triggers = [k for k in pressed if k not in self._MODIFIERS]
                if triggers:
                    mods   = [k for k in pressed if k in self._MODIFIERS]
                    hotkey = '+'.join(mods + triggers)
                    done.set()
                    try: keyboard.unhook(hook_ref[0])
                    except Exception: pass
                    hook_ref[0] = None
                    self.captured.emit(hotkey)
                else:
                    if name in pressed:
                        pressed.remove(name)

        try:
            hook_ref[0] = keyboard.hook(on_event)
            done.wait(timeout=30)
        except Exception as e:
            print(f"[HotkeyCaptureThread] error: {e}")
        finally:
            if hook_ref[0] is not None:
                try: keyboard.unhook(hook_ref[0])
                except Exception: pass


# ============================================================
# hotkey キャプチャオーバーレイ
# ============================================================
class HotkeyCaptureOverlay(QWidget):
    def __init__(self, action_name, callback, parent=None):
        super().__init__(parent, Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint | Qt.Window)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.callback = callback
        self._thread  = None

        layout = QVBoxLayout(self)
        layout.setContentsMargins(36, 24, 36, 24)
        layout.setSpacing(14)

        title = QLabel(f"「{action_name}」")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("color: #ffdd55; font-size: 15px; font-weight: bold; background: transparent;")

        msg = QLabel("キーを押してください\n（組み合わせキーも対応）")
        msg.setAlignment(Qt.AlignCenter)
        msg.setStyleSheet("color: white; font-size: 13px; background: transparent;")

        btn_style = ("QPushButton { color: white; background: rgba(80,80,80,200); border: 1px solid #888;"
                     " border-radius: 4px; padding: 4px 16px; font-size: 12px; }"
                     "QPushButton:hover { background: rgba(120,120,120,220); }")
        cancel_btn = QPushButton("キャンセル")
        cancel_btn.setStyleSheet(btn_style)
        cancel_btn.clicked.connect(self.cancel)

        layout.addWidget(title)
        layout.addWidget(msg)
        layout.addWidget(cancel_btn, alignment=Qt.AlignCenter)

        self.setStyleSheet("HotkeyCaptureOverlay { background: rgba(30,30,30,210);"
                           " border: 1px solid #666; border-radius: 10px; }")
        self.adjustSize()
        _, center = _cursor_screen_center()
        self.move(center - self.rect().center())
        self._start_capture()

    def _start_capture(self):
        self._thread = HotkeyCaptureThread()
        self._thread.captured.connect(self._on_captured)
        self._thread.start()

    def _on_captured(self, key):
        self.callback(key)
        self.close()

    def cancel(self):
        if self._thread and self._thread.isRunning():
            self._thread.terminate(); self._thread.wait()
        self.close()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape:
            self.cancel()


# ============================================================
# 汎用ボタンダイアログ基底
# ============================================================
class _ButtonDialog(QWidget):
    _STYLE = ("QPushButton { color: white; background: rgba(60,60,60,200); border: 1px solid #888;"
              " border-radius: 4px; padding: 6px 20px; font-size: 13px; }"
              "QPushButton:hover { background: rgba(100,100,100,220); }")

    def __init__(self, title_text, buttons, parent=None):
        super().__init__(parent, Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint | Qt.Window)
        self.setAttribute(Qt.WA_TranslucentBackground)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(36, 24, 36, 24)
        layout.setSpacing(14)

        title = QLabel(title_text)
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("color: #ffdd55; font-size: 15px; font-weight: bold; background: transparent;")
        layout.addWidget(title)

        for label, cb in buttons:
            btn = QPushButton(label)
            btn.setStyleSheet(self._STYLE)
            if cb:
                btn.clicked.connect(lambda checked, f=cb: (f(), self.close()))
            else:
                btn.clicked.connect(self.close)
            layout.addWidget(btn)

        self.setStyleSheet(f"{type(self).__name__} {{ background: rgba(30,30,30,210);"
                           " border: 1px solid #666; border-radius: 10px; }")
        self.adjustSize()
        _, center = _cursor_screen_center()
        self.move(center - self.rect().center())

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape:
            self.close()


class ResetPositionDialog(_ButtonDialog):
    def __init__(self, on_center, on_config, parent=None):
        super().__init__("位置をリセット", [
            ("モニタ中央へ移動", on_center),
            ("設定位置に戻す",   on_config),
            ("キャンセル",       None),
        ], parent)


class LayoutSelectDialog(_ButtonDialog):
    def __init__(self, on_layout1, on_layout2, parent=None):
        super().__init__("どちらに保存しますか？", [
            ("配置1に保存", on_layout1),
            ("配置2に保存", on_layout2),
            ("キャンセル",  None),
        ], parent)


# ============================================================
# 縁取り文字デリゲート（テーブル用）
# ============================================================
class OutlineTextDelegate(QStyledItemDelegate):
    def __init__(self, default_font_size, outline_color, parent=None):
        super().__init__(parent)
        self.default_font_size = default_font_size
        self.outline_color = QColor(outline_color)

    def paint(self, painter, option, index):
        text = index.data(Qt.DisplayRole)
        if not text:
            return

        custom_size = index.data(Qt.UserRole)

        # 1行目の見出しセルは固定サイズで描画
        if index.row() == 0:
            current_font_size = self.default_font_size
        else:
            current_font_size = int(custom_size) if custom_size is not None else self.default_font_size

        color_data = index.data(Qt.ForegroundRole)
        font_color = QColor(color_data) if color_data else QColor("white")

        painter.save()
        painter.setRenderHint(QPainter.Antialiasing)
        painter.setRenderHint(QPainter.TextAntialiasing)

        font = QFont('Meiryo', current_font_size, QFont.Bold)
        painter.setFont(font)

        rect = option.rect
        metrics = QFontMetrics(font)
        x = rect.left() + 5
        y = rect.top() + (rect.height() + metrics.ascent() - metrics.descent()) // 2

        path = QPainterPath()
        path.addText(x, y, font, text)

        pen = QPen(self.outline_color)
        pen.setWidthF(4.0)
        pen.setJoinStyle(Qt.RoundJoin)
        painter.setPen(pen)
        painter.drawPath(path)
        painter.fillPath(path, QBrush(font_color))
        painter.restore()


# ============================================================
# ドラッグ移動テーブル
# ============================================================
class DraggableTableWidget(QTableWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_window = parent
        self._drag_start = None

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self._drag_start = event.globalPos(); event.accept()
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if event.buttons() == Qt.LeftButton and self._drag_start is not None:
            delta = event.globalPos() - self._drag_start
            self._drag_start = event.globalPos()
            self.parent_window.move(self.parent_window.pos() + delta); event.accept()
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        if event.button() == Qt.LeftButton:
            self._drag_start = None
        super().mouseReleaseEvent(event)


# ============================================================
# キーボードイベント中継用シグナラー
# ============================================================
class KeyboardSignaler(QObject):
    switch_requested         = pyqtSignal()
    preset_switch_requested  = pyqtSignal()
    reset_position_requested = pyqtSignal()


class TimerSignaler(QObject):
    do_start_stop  = pyqtSignal()
    do_log_a       = pyqtSignal()
    do_log_ex      = pyqtSignal()
    do_log_loop    = pyqtSignal()
    do_log_loopB   = pyqtSignal()
    do_increment   = pyqtSignal()
    do_decrement   = pyqtSignal()
    do_ai_count_2  = pyqtSignal()
    do_ai_count_3  = pyqtSignal()


# ============================================================
# プリセットローダー
# ============================================================
def load_presets(config_path):
    config = configparser.ConfigParser()
    config.read(config_path, encoding='utf-8')
    presets = []
    i = 1
    while True:
        section = f'Preset{i}'
        if not config.has_section(section): break
        raw_file = config.get(section, 'table_file', fallback='data.xlsx')
        p = {
            'name':              config.get(section, 'name', fallback=section),
            'table_file':        xlsx_path(raw_file),
            'initial_time':      config.get(section, 'initial_time',         fallback='10:00'),
            'activation_normal': config.getint(section, 'activation_normal', fallback=0) / 1000,
            'activation_loopA':  config.getint(section, 'activation_loopA',  fallback=0) / 1000,
            'activation_loopB':  config.getint(section, 'activation_loopB',  fallback=0) / 1000,
            'activation_ex':     config.getint(section, 'activation_ex',     fallback=0) / 1000,
            'loopA_time':   [int(t.strip()) / 1000 for t in config.get(section, 'loopA_time',   fallback='1000').split(',')],
            'loopB_time':   [int(t.strip()) / 1000 for t in config.get(section, 'loopB_time',   fallback='1000').split(',')],
            'log_label':    config.get(section, 'log_label',    fallback='').replace('\\n', '\n'),
            'log_label_ex': config.get(section, 'log_label_ex', fallback='').replace('\\n', '\n'),
            'loopA_labels':    [l.strip().replace('\\n', '\n') for l in config.get(section, 'loopA_labels', fallback='').split(',')],
            'loopB_labels':    [l.strip().replace('\\n', '\n') for l in config.get(section, 'loopB_labels', fallback='').split(',')],
            'loopA_stop_time': config.getint(section, 'loopA_stop_time', fallback=0) / 1000,
            'loopB_stop_time': config.getint(section, 'loopB_stop_time', fallback=0) / 1000,
        }
        presets.append(p); i += 1
    return presets


# ============================================================
# Excel 読み込みスレッド
# ============================================================
class ExcelLoaderThread(QThread):
    loaded = pyqtSignal(dict)

    def __init__(self, file_path, sheet_index):
        super().__init__()
        self.file_path   = file_path
        self.sheet_index = sheet_index

    def run(self):
        result = {'rows_data': [], 'col_widths': [], 'row_heights': {},
                  'row_count': 0, 'col_count': 0, 'sheet_index': self.sheet_index}
        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            sheet_names = wb.sheetnames
            if not sheet_names:
                self.loaded.emit(result); return
            idx = self.sheet_index % len(sheet_names)
            result['sheet_index'] = idx
            ws   = wb[sheet_names[idx]]
            rows = list(ws.rows)
            if not rows:
                self.loaded.emit(result); wb.close(); return
            row_count = len(rows)
            col_count = max(len(row) for row in rows)
            result['row_count'] = row_count
            result['col_count'] = col_count
            for j in range(col_count):
                col_letter = openpyxl.utils.get_column_letter(j + 1)
                if col_letter in ws.column_dimensions:
                    w = ws.column_dimensions[col_letter].width
                    result['col_widths'].append(int(w * 8) if w else 100)
                else:
                    result['col_widths'].append(100)
            for i in range(row_count):
                if i + 1 in ws.row_dimensions:
                    h = ws.row_dimensions[i + 1].height
                    if h: result['row_heights'][i] = int(h * 1.3)
            for i, row in enumerate(rows):
                row_data = []
                for j, cell in enumerate(row):

                    val = cell.value
                    if isinstance(val, float) and val.is_integer(): val = int(val)
                    text = str(val) if val is not None else ""
                    font_color = "white"; font_size = None
                    if cell.font:
                        try:
                            c = cell.font.color
                            if getattr(c, 'type', None) == 'rgb':
                                color_val = c.rgb
                                if (isinstance(color_val, str) and len(color_val) == 8
                                        and color_val[:2] != '00'
                                        and color_val.upper() != 'FF000000'):
                                    font_color = f"#{color_val[2:]}"
                        except: pass
                        sz = cell.font.sz
                        if sz and float(sz) >= 6.0:
                            font_size = sz
                    row_data.append({'text': text, 'font_color': font_color, 'font_size': font_size})
                result['rows_data'].append(row_data)
            wb.close()
        except Exception as e:
            print(f"ExcelLoaderThread Error: {e}")
        self.loaded.emit(result)


# ============================================================
# Excel 表示ウィジェット
# ============================================================
class ExcelTableWindow(QWidget):
    _CONFIG_KEY_MAP = {
        'switch_key':         'sheet_switch_key',
        'preset_switch_key':  'preset_switch_key',
        'reset_position_key': 'reset_position_key',
    }

    def __init__(self, config_path, presets, parent=None):
        super().__init__(parent)
        self.config_path          = config_path
        self._drag_start          = None
        self.current_sheet_index  = 0
        self.signaler             = KeyboardSignaler()
        self.presets              = presets
        self.current_preset_index = 0
        self._hotkey_handler      = None
        self._loading             = False
        self._preset_lock         = threading.Lock()
        self._loader_thread       = None
        self.timer_win            = None
        self.main_win             = None
        self._capture_overlay     = None
        self._reset_dialog        = None
        self._layout_dialog       = None
        self.current_layout       = 1
        self.refresh_timer        = QTimer(self)
        self.refresh_timer.timeout.connect(self.maintain_system)
        self.refresh_timer.start(30000)
        self.initUI()

    def initUI(self):
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setAutoFillBackground(False)

        config = configparser.ConfigParser()
        config.read(self.config_path, encoding='utf-8')

        self.f_size        = config.getint('Settings', 'label_font_size',    fallback=11)
        self.o_color       = config.get('Settings',    'outline_color',      fallback='#101010')
        self.switch_key         = config.get('Settings', 'sheet_switch_key',    fallback='ctrl+tab')
        self.preset_switch_key  = config.get('Settings', 'preset_switch_key',   fallback='ctrl+shift+tab')
        self.reset_position_key = config.get('Settings', 'reset_position_key',  fallback='ctrl+home')
        self.table_align   = config.get('Settings', 'table_align', fallback='right').strip().lower()
        if self.table_align not in ('left', 'right'):
            self.table_align = 'right'
        self.table_visible = config.getboolean('Settings', 'table_visible', fallback=True)

        self.tx = [0, config.getint('Settings', 'table_x',   fallback=1920),
                      config.getint('Settings', 'table_x_2', fallback=1920)]
        self.ty = [0, config.getint('Settings', 'table_y',   fallback=0),
                      config.getint('Settings', 'table_y_2', fallback=0)]

        if self.presets:
            self.file_path = self.presets[self.current_preset_index]['table_file']
        else:
            raw = config.get('Settings', 'table_file', fallback='data.xlsx')
            self.file_path = xlsx_path(raw)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setAlignment(Qt.AlignRight | Qt.AlignTop)

        self.table = DraggableTableWidget(self)
        self.table.setStyleSheet("background: transparent; border: none; gridline-color: transparent;")
        self.table.setFrameShape(QTableWidget.NoFrame)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.table.horizontalHeader().setVisible(False)
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionMode(QTableWidget.NoSelection)
        self.table.setFocusPolicy(Qt.NoFocus)

        self.delegate = OutlineTextDelegate(self.f_size, self.o_color)
        self.table.setItemDelegate(self.delegate)
        layout.addWidget(self.table)
        self.resize(100, 100)

        self.signaler.switch_requested.connect(self.next_sheet)
        self.signaler.preset_switch_requested.connect(self.next_preset)
        self.signaler.reset_position_requested.connect(self.show_reset_position_dialog)
        self.setup_hotkey()
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self.showContextMenu)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.showContextMenu)

        self._position_initialized = False
        self._start_load(self.current_sheet_index)
        self.table.setVisible(self.table_visible)

    def _apply_table_position(self):
        if self._position_initialized: return
        parent = self.parent()
        ox = parent.geometry().x() if parent else 0
        oy = parent.geometry().y() if parent else 0
        tx = self.tx[self.current_layout]
        ty = self.ty[self.current_layout]
        x = (tx - ox - self.width()) if self.table_align == 'right' else (tx - ox)
        self.move(x, ty - oy)
        self._position_initialized = True

    def _origin(self):
        parent = self.parent()
        return (parent.geometry().x() if parent else 0,
                parent.geometry().y() if parent else 0)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self._drag_start = event.globalPos(); event.accept()

    def mouseMoveEvent(self, event):
        if event.buttons() == Qt.LeftButton and self._drag_start is not None:
            delta = event.globalPos() - self._drag_start
            self._drag_start = event.globalPos()
            self.move(self.pos() + delta); event.accept()

    def mouseReleaseEvent(self, event):
        if event.button() == Qt.LeftButton:
            self._drag_start = None

    @staticmethod
    def _parse_hotkey(hotkey_str):
        parts = [p.strip().lower() for p in hotkey_str.split('+')]
        return parts[-1], parts[:-1]

    def _on_key_event(self, event):
        if event.event_type != keyboard.KEY_DOWN: return
        if not event.name: return
        name = HotkeyCaptureThread._norm(event.name.lower())
        trigger, mods = self._parse_hotkey(self.switch_key)
        if name == trigger and all(keyboard.is_pressed(m) for m in mods):
            self.signaler.switch_requested.emit(); return
        if self.presets:
            t2, m2 = self._parse_hotkey(self.preset_switch_key)
            if name == t2 and all(keyboard.is_pressed(m) for m in m2):
                self.signaler.preset_switch_requested.emit(); return
        t3, m3 = self._parse_hotkey(self.reset_position_key)
        if name == t3 and all(keyboard.is_pressed(m) for m in m3):
            self.signaler.reset_position_requested.emit()

    def setup_hotkey(self):
        if self._hotkey_handler is not None:
            try: keyboard.unhook(self._hotkey_handler)
            except Exception: pass
            self._hotkey_handler = None
        self._hotkey_handler = keyboard.hook(self._on_key_event)

    def set_hotkey(self, attr, value):
        setattr(self, attr, value)
        config_key = self._CONFIG_KEY_MAP.get(attr, attr)
        safe_set_config(self.config_path, config_key, value)
        self.setup_hotkey()

    def maintain_system(self):
        gc.collect(); self.setup_hotkey()

    def _start_load(self, sheet_index):
        if self._loading: return
        self._loading = True
        if self._loader_thread is not None:
            try: self._loader_thread.loaded.disconnect()
            except Exception: pass
            if self._loader_thread.isRunning():
                self._loader_thread.quit(); self._loader_thread.wait()
        self._loader_thread = ExcelLoaderThread(self.file_path, sheet_index)
        self._loader_thread.loaded.connect(self._apply_sheet_data)
        self._loader_thread.start()
        QTimer.singleShot(10000, self._loading_timeout)

    def _loading_timeout(self):
        if self._loading:
            print("[warn] _loading timeout: force released")
            self._loading = False

    def _apply_sheet_data(self, data):
        self._loading = False
        self.current_sheet_index = data['sheet_index']
        if not data['rows_data']:
            self.table.setRowCount(0); return
        self.table.setUpdatesEnabled(False)
        self.table.setRowCount(data['row_count'])
        self.table.setColumnCount(data['col_count'])
        for j, w in enumerate(data['col_widths']): self.table.setColumnWidth(j, w)
        for i, h in data['row_heights'].items():   self.table.setRowHeight(i, h)
        for i, row_data in enumerate(data['rows_data']):
            for j, cell_data in enumerate(row_data):
                item = QTableWidgetItem(cell_data['text'])
                item.setFlags(Qt.ItemIsEnabled)
                item.setForeground(QBrush(QColor(cell_data['font_color'])))
                if cell_data['font_size'] is not None:
                    item.setData(Qt.UserRole, cell_data['font_size'])
                self.table.setItem(i, j, item)
        t_width  = self.table.horizontalHeader().length() + 4
        t_height = self.table.verticalHeader().length() + 4
        old_anchor = (self.x() + self.width()) if self.table_align == 'right' else self.x()
        self.table.setFixedSize(t_width, t_height)
        self.setFixedSize(t_width, t_height)
        if not self._position_initialized:
            self._apply_table_position()
        else:
            self.move((old_anchor - t_width) if self.table_align == 'right' else old_anchor, self.y())
        self.table.setUpdatesEnabled(True)

    def next_sheet(self):
        self.current_sheet_index += 1
        self._start_load(self.current_sheet_index)

    def set_timer_window(self, timer_win): self.timer_win = timer_win
    def set_main_window(self, main_win):   self.main_win  = main_win

    def next_preset(self):
        if not self.presets: return
        self.select_preset((self.current_preset_index + 1) % len(self.presets))

    def select_preset(self, index):
        if not self.presets: return
        if not self._preset_lock.acquire(blocking=False): return
        try:
            self.current_preset_index = index
            preset = self.presets[self.current_preset_index]
            self.file_path = preset['table_file']
            self.current_sheet_index = 0
            if self.timer_win: self.timer_win.apply_preset(preset)
            self._start_load(self.current_sheet_index)
        finally:
            self._preset_lock.release()

    def switch_layout(self, n):
        self.current_layout = n
        ox, oy = self._origin()
        tx = self.tx[n]; ty = self.ty[n]
        x = (tx - ox - self.width()) if self.table_align == 'right' else (tx - ox)
        self.move(x, ty - oy)
        if self.timer_win:
            wx = self.timer_win.timer_x_list[n]
            wy = self.timer_win.timer_y_list[n]
            self.timer_win.move(wx - ox, wy - oy)

    def show_save_position_dialog(self):
        if self._layout_dialog is not None:
            try: self._layout_dialog.close()
            except Exception: pass
        self._layout_dialog = LayoutSelectDialog(
            on_layout1=lambda: self._save_positions_to(1),
            on_layout2=lambda: self._save_positions_to(2),
        )
        self._layout_dialog.show()

    def _save_positions_to(self, n):
        ox, oy = self._origin()
        new_tx = (ox + self.x() + self.width()) if self.table_align == 'right' else (ox + self.x())
        new_ty = oy + self.y()
        key_x = 'table_x' if n == 1 else 'table_x_2'
        key_y = 'table_y' if n == 1 else 'table_y_2'
        safe_set_config(self.config_path, key_x, str(new_tx))
        safe_set_config(self.config_path, key_y, str(new_ty))
        self.tx[n] = new_tx
        self.ty[n] = new_ty
        if self.timer_win:
            self.timer_win.save_position_to(n, ox, oy)
        self.current_layout = n

    def save_position(self):
        self.show_save_position_dialog()

    def set_table_align(self, align):
        if align not in ('left', 'right') or align == self.table_align: return
        ox, oy = self._origin()
        anchor = (ox + self.x() + self.width()) if self.table_align == 'right' else (ox + self.x())
        self.table_align = align
        safe_set_config(self.config_path, 'table_align', align)
        new_sx = (anchor - self.width()) if align == 'right' else anchor
        new_sy = oy + self.y()
        desktop = QApplication.desktop()
        sr = desktop.screenGeometry(desktop.screenNumber(self))
        new_sx = max(sr.left(), min(new_sx, sr.right()  - self.width()))
        new_sy = max(sr.top(),  min(new_sy, sr.bottom() - self.height()))
        self.move(new_sx - ox, new_sy - oy)

    def set_table_visible(self, visible):
        self.table_visible = visible
        self.table.setVisible(visible)
        safe_set_config(self.config_path, 'table_visible', '1' if visible else '0')

    def reload_from_config(self):
        config = configparser.ConfigParser()
        config.read(self.config_path, encoding='utf-8')
        self.switch_key         = config.get('Settings', 'sheet_switch_key',    fallback=self.switch_key)
        self.preset_switch_key  = config.get('Settings', 'preset_switch_key',   fallback=self.preset_switch_key)
        self.reset_position_key = config.get('Settings', 'reset_position_key',  fallback=self.reset_position_key)
        self.table_align        = config.get('Settings', 'table_align',         fallback=self.table_align).strip().lower()
        self.table_visible      = config.getboolean('Settings', 'table_visible', fallback=self.table_visible)
        self.tx = [0, config.getint('Settings', 'table_x',   fallback=self.tx[1]),
                      config.getint('Settings', 'table_x_2', fallback=self.tx[2])]
        self.ty = [0, config.getint('Settings', 'table_y',   fallback=self.ty[1]),
                      config.getint('Settings', 'table_y_2', fallback=self.ty[2])]
        self.table.setVisible(self.table_visible)
        self.setup_hotkey()

    def show_restore_config_dialog(self):
        bak = self.config_path + '.bak'
        if not os.path.isfile(bak): return
        if restore_config(self.config_path):
            self.reload_from_config()
            if self.timer_win:
                self.timer_win.reload_from_config()
            if self.main_win:
                cfg = configparser.ConfigParser()
                cfg.read(self.config_path, encoding='utf-8')
                self.main_win.set_opacity(cfg.getint('Settings', 'opacity', fallback=self.main_win._opacity))

    def show_reset_position_dialog(self):
        if self._reset_dialog is not None:
            try: self._reset_dialog.close()
            except Exception: pass
        self._reset_dialog = ResetPositionDialog(
            on_center=self._reset_to_center,
            on_config=self._reset_to_config,
        )
        self._reset_dialog.show()

    def _reset_to_center(self):
        _, center = _cursor_screen_center()
        ox, oy = self._origin()
        if self.timer_win:
            tw, th = self.timer_win.width(), self.timer_win.height()
            self.timer_win.move(center.x() - tw // 2 - ox, center.y() - th // 2 - oy)
        tw2, th2 = self.width(), self.height()
        self.move(center.x() - tw2 // 2 - ox, center.y() - th2 // 2 - oy)

    def _reset_to_config(self):
        self.switch_layout(self.current_layout)

    def _start_hotkey_capture(self, action_name, attr, target):
        if self._capture_overlay is not None:
            try: self._capture_overlay.cancel()
            except Exception: pass
        self._capture_overlay = HotkeyCaptureOverlay(
            action_name,
            lambda key, a=attr, t=target: t.set_hotkey(a, key),
            parent=None)
        self._capture_overlay.show()

    def _build_color_menu(self, parent_menu, tmr):
        if tmr is None: return
        color_menu = parent_menu.addMenu("フォントカラーの設定")
        for label, attr in [("通常時",       "timer_font_color"),
                             ("動作中",       "timer_font_color_while_running"),
                             ("EX",          "timer_font_color_ex"),
                             ("AIカウント",   "ai_count_font_color"),
                             ("縁取り",       "outline_color")]:
            cur = getattr(tmr, attr, '#ffffff')
            color_menu.addAction(_make_color_icon(cur), f"{label}  {cur}").triggered.connect(
                lambda checked, a=attr, t=tmr: t.open_color_dialog(a))

    def _build_settings_save_menu(self, menu):
        twr = self; tmr = self.timer_win
        sp  = menu.addMenu("設定・保存")

        bak_exists = os.path.isfile(self.config_path + '.bak')
        act_restore = sp.addAction("設定を戻す")
        act_restore.triggered.connect(self.show_restore_config_dialog)
        act_restore.setEnabled(bak_exists)
        sp.addSeparator()

        sp.addAction("現在の位置を保存").triggered.connect(self.show_save_position_dialog)

        if self.main_win:
            cur_op = self.main_win._opacity
            op_menu = sp.addMenu("透過率の設定")
            for v in range(100, 0, -10):
                lbl = f"{'▶ ' if v == cur_op else '　'}{v}%"
                op_menu.addAction(lbl).triggered.connect(
                    lambda checked, val=v: self.main_win.set_opacity(val))

        if tmr:
            # ---- AIカウント ----
            ai_menu = sp.addMenu("AIカウントの表示設定")
            vis_ai = tmr.ai_count_visible
            ai_menu.addAction(f"{'▶ ' if vis_ai     else '　'}表示").triggered.connect(
                lambda: tmr.set_ai_count_visible(True))
            ai_menu.addAction(f"{'▶ ' if not vis_ai else '　'}非表示").triggered.connect(
                lambda: tmr.set_ai_count_visible(False))

        xlsx_menu = sp.addMenu("テーブルの表示設定")
        for align, label in (('right', '右揃え'), ('left', '左揃え')):
            lbl = f"{'▶ ' if twr.table_align == align else '　'}{label}"
            xlsx_menu.addAction(lbl).triggered.connect(
                lambda checked, a=align: twr.set_table_align(a))
        xlsx_menu.addSeparator()
        vis = getattr(twr, 'table_visible', True)
        xlsx_menu.addAction(f"{'▶ ' if vis     else '　'}表示").triggered.connect(
            lambda: twr.set_table_visible(True))
        xlsx_menu.addAction(f"{'▶ ' if not vis else '　'}非表示").triggered.connect(
            lambda: twr.set_table_visible(False))

        if tmr:
            font_menu = sp.addMenu("フォントサイズの設定")
            timer_font_menu = font_menu.addMenu("タイマー")
            for fs in range(8, 81, 4):
                lbl = f"{'▶ ' if fs == tmr.timer_font_size else '　'}{fs}pt"
                timer_font_menu.addAction(lbl).triggered.connect(
                    lambda checked, v=fs: tmr.set_font_size(v))
            label_font_menu = font_menu.addMenu("ラベル")
            for fs in range(6, 33, 2):
                lbl = f"{'▶ ' if fs == tmr.label_font_size else '　'}{fs}pt"
                label_font_menu.addAction(lbl).triggered.connect(
                    lambda checked, v=fs: tmr.set_label_font_size(v))

            self._build_color_menu(sp, tmr)

            vol_menu = sp.addMenu("音量の設定")
            for v in range(100, -1, -10):
                lbl = f"{'▶ ' if v == tmr.sound_volume else '　'}{v}%"
                vol_menu.addAction(lbl).triggered.connect(
                    lambda checked, val=v: tmr.set_sound_volume(val))

        hk_menu = sp.addMenu("Hotkeyの設定")
        hotkey_items = [
            ("タイマー開始/停止", "start_key",         tmr),
            ("タイマー記録",            "log_key",            tmr),
            ("タイマー記録[EX表示]",           "log_key_ex",         tmr),
            ("タイマー記録[ループA開始]",       "log_key_loop",       tmr),
            ("タイマー記録[ループB開始]",       "log_key_loopB",      tmr),
            ("タイマー ズレ修正[増加]",    "increment_key",      tmr),
            ("タイマー ズレ修正[減少]",    "decrement_key",      tmr),
            ("シート切替",        "switch_key",         twr),
            ("プリセット切替",    "preset_switch_key",  twr),
            ("位置リセット",      "reset_position_key", twr),
        ]
        for name, attr, target in hotkey_items:
            if target is None: continue
            cur = getattr(target, attr, "") or "未設定"
            hk_menu.addAction(f"{name}  [{cur}]").triggered.connect(
                lambda checked, n=name, a=attr, t=target: self._start_hotkey_capture(n, a, t))
        if tmr:
            hk_menu.addSeparator()
            for ai_name, ai_attr in [("AIカウント (1-2)",   "ai_count_key_2"),
                                      ("AIカウント (1-2-3)", "ai_count_key_3")]:
                ai_cur = getattr(tmr, ai_attr, "") or "未設定"
                hk_menu.addAction(f"{ai_name}  [{ai_cur}]").triggered.connect(
                    lambda checked, n=ai_name, a=ai_attr, t=tmr: self._start_hotkey_capture(n, a, t))

    def _add_common_top_items(self, menu):
        menu.addAction("位置をリセット").triggered.connect(self.show_reset_position_dialog)
        layout_menu = menu.addMenu("配置切替")
        for n in (1, 2):
            lbl = f"{'▶ ' if n == self.current_layout else '　'}配置{n}"
            layout_menu.addAction(lbl).triggered.connect(
                lambda checked, v=n: self.switch_layout(v))
        if self.presets:
            preset_menu = menu.addMenu("プリセット切替")
            for i, p in enumerate(self.presets):
                lbl = f"{'▶ ' if i == self.current_preset_index else '　'}{p['name']}"
                preset_menu.addAction(lbl).triggered.connect(
                    lambda checked, idx=i: self.select_preset(idx))
            menu.addSeparator()

    def showContextMenu(self, pos):
        global_pos = self.sender().mapToGlobal(pos) if self.sender() else self.mapToGlobal(pos)
        menu = QMenu(self)
        menu.setStyleSheet("QMenu { color: white; background-color: rgba(50, 50, 50, 150); border: 1px solid #555; }"
                           "QMenu::item:selected { background-color: rgba(100, 100, 100, 200); }")
        self._add_common_top_items(menu)
        self._build_settings_save_menu(menu)
        menu.addSeparator()
        menu.addAction("閉じる").triggered.connect(QApplication.instance().quit)
        menu.addAction("キャンセル")
        menu.exec_(global_pos)


# ============================================================
# タイマースレッド
# ============================================================
class TimerThread(QThread):
    timer_updated = pyqtSignal(float)

    def __init__(self, initial_time_seconds, reset_callback):
        super().__init__()
        self.initial_time_seconds = initial_time_seconds
        self.current_time         = initial_time_seconds
        self.running              = False
        self.reset_callback       = reset_callback

    def run(self):
        try:
            self.running = True
            start_time = time.time()
            while self.running and self.current_time > 0:
                elapsed = time.time() - start_time
                self.current_time = self.initial_time_seconds - elapsed
                if self.current_time < 0: self.current_time = 0
                self.timer_updated.emit(self.current_time)
                time.sleep(0.02)
            if self.current_time <= 0 and self.running:
                self.running = False; self.reset_callback()
        except Exception as e:
            print(f"[TimerThread] 予期しないエラー: {e}")

    def stop(self): self.running = False


# ============================================================
# タイマーウィジェット
# ============================================================
class TransparentWindow(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.timer_thread          = None
        self.last_played_target    = -1
        self._drag_start           = None
        self.table_win             = None
        self.main_win              = None
        self._pixmap_cache         = {}
        self._movie_cache          = {}
        self._gif_labels           = []
        self._current_label_str    = ""
        self._cached_segments      = []
        self._cached_ex_segments   = []
        self._key_hooks            = []
        self._capture_overlay      = None
        self.initUI()

    def set_table_window(self, table_win): self.table_win = table_win
    def set_main_window(self, main_win):   self.main_win  = main_win

    def initUI(self):
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setAutoFillBackground(False)
        self.loadConfig()
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0); layout.setSpacing(0)
        self.setLayout(layout)
        self.setFixedSize(self.fixed_width, self.fixed_height)
        self.setupUIElements(layout)
        self.setupEvents()
        self.setupLogs()
        self._apply_initial_position()

    def _apply_initial_position(self):
        parent = self.parent()
        ox = parent.geometry().x() if parent else 0
        oy = parent.geometry().y() if parent else 0
        self.move(self.timer_x_list[1] - ox, self.timer_y_list[1] - oy)

    def loadConfig(self):
        config = configparser.ConfigParser()
        config.read('config.ini', encoding='utf-8')
        # ---- 色・フォント ----
        self.timer_font_color               = config.get('Settings', 'timer_font_color',               fallback='#ffffff')
        self.timer_font_color_while_running = config.get('Settings', 'timer_font_color_while_running', fallback='#ff4444')
        self.timer_font_color_ex            = config.get('Settings', 'timer_font_color_ex',            fallback='#ffffff')
        self.outline_color                  = config.get('Settings', 'outline_color',                  fallback='#101010')
        self.timer_font_size                = config.getint('Settings', 'timer_font_size',              fallback=48)
        self.label_font_size                = config.getint('Settings', 'label_font_size',              fallback=8)
        # ---- タイマー設定 ----
        self.initial_time = self.reset_time = config.get('Settings', 'initial_time', fallback='10:00')
        self.fixed_width  = 1280
        self.fixed_height = 720
        # ---- hotkey ----
        self.start_key     = config.get('Settings', 'start_key',     fallback='F1')
        self.log_key       = config.get('Settings', 'log_key',       fallback='F2')
        self.log_key_ex    = config.get('Settings', 'log_key_ex',    fallback='')
        self.log_key_loop  = config.get('Settings', 'log_key_loop',  fallback='F3')
        self.log_key_loopB = config.get('Settings', 'log_key_loopB', fallback='')
        self.increment_key = config.get('Settings', 'increment_key', fallback='F4')
        self.decrement_key = config.get('Settings', 'decrement_key', fallback='F5')
        # ---- AIカウント ----
        self.ai_count_value      = 1
        self.ai_count_visible    = config.getboolean('Settings', 'ai_count_visible',    fallback=True)
        self.ai_count_font_color = config.get('Settings',        'ai_count_font_color', fallback='#ffffff')
        self.ai_count_key_2      = config.get('Settings',        'ai_count_key_2',      fallback='')
        self.ai_count_key_3      = config.get('Settings',        'ai_count_key_3',      fallback='')
        # ---- サウンド ----
        self.play_start_sound  = config.getboolean('Settings', 'play_start_sound',  fallback=False)
        self.play_bomb_sound   = config.getboolean('Settings', 'play_activation_sound', fallback=False)
        self.play_record_sound = config.getboolean('Settings', 'play_record_sound', fallback=False)
        self.sound_delay       = config.getint('Settings', 'sound_delay',  fallback=0) / 1000
        self.sound_volume      = config.getint('Settings', 'sound_volume', fallback=50)
        self.image_label_height = config.getint('Settings', 'image_label_height', fallback=0)
        # ---- 配置1・配置2 ----
        self.timer_x_list = [0,
            config.getint('Settings', 'timer_x',   fallback=0),
            config.getint('Settings', 'timer_x_2', fallback=0)]
        self.timer_y_list = [0,
            config.getint('Settings', 'timer_y',   fallback=0),
            config.getint('Settings', 'timer_y_2', fallback=0)]

        self._presets = load_presets('config.ini')
        if self._presets:
            self.apply_preset(self._presets[0], reset_logs=False)
        else:
            self.activation_normal = config.getint('Settings', 'activation_normal', fallback=0) / 1000
            self.activation_loopA  = config.getint('Settings', 'activation_loopA',  fallback=0) / 1000
            self.activation_loopB  = config.getint('Settings', 'activation_loopB',  fallback=0) / 1000
            self.activation_ex     = config.getint('Settings', 'activation_ex',     fallback=0) / 1000
            self.loopA_times  = [int(t.strip()) / 1000 for t in config.get('Settings', 'loopA_time', fallback='1000').split(',')]
            self.loopA_labels = [l.strip().replace('\\n', '\n') for l in config.get('Settings', 'loopA_labels', fallback='').split(',')]
            self.loopB_times  = [int(t.strip()) / 1000 for t in config.get('Settings', 'loopB_time', fallback='1000').split(',')]
            self.loopB_labels    = [l.strip().replace('\\n', '\n') for l in config.get('Settings', 'loopB_labels', fallback='').split(',')]
            self.log_label       = config.get('Settings', 'log_label',    fallback='').replace('\\n', '\n')
            self.log_label_ex    = config.get('Settings', 'log_label_ex', fallback='').replace('\\n', '\n')
            self.loopA_stop_time = 0
            self.loopB_stop_time = 0
        self.current_loopA_index = 0
        self.current_loopB_index = 0

    def apply_preset(self, preset, reset_logs=True):
        self.activation_normal = preset['activation_normal']
        self.activation_loopA  = preset['activation_loopA']
        self.activation_loopB  = preset['activation_loopB']
        self.activation_ex     = preset['activation_ex']
        self.loopA_times       = preset['loopA_time']
        self.loopB_times       = preset['loopB_time']
        self.log_label         = preset['log_label']
        self.log_label_ex      = preset['log_label_ex']
        self.loopA_labels      = preset['loopA_labels']
        self.loopB_labels      = preset['loopB_labels']
        self.loopA_stop_time   = preset['loopA_stop_time']
        self.loopB_stop_time   = preset['loopB_stop_time']
        self.initial_time      = preset['initial_time']
        self.reset_time        = preset['initial_time']
        if reset_logs:
            self.clear_logs_only()
            self.update_display_text(self.initial_time, "")

    def setupUIElements(self, layout):
        self.setupMediaPlayer()
        self.label = QLabel(f"{self.initial_time}\n ")
        self.label.setAlignment(Qt.AlignCenter)
        self.label.setStyleSheet("color: transparent; background: transparent;")
        self.label.setAttribute(Qt.WA_TransparentForMouseEvents, True)
        layout.addWidget(self.label)
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self.showContextMenu)

    def _scale_pixmap(self, pm):
        if self.image_label_height > 0 and not pm.isNull() and pm.height() > 0:
            return pm.scaledToHeight(self.image_label_height, Qt.SmoothTransformation)
        return pm

    def _get_pixmap(self, filename):
        if filename not in self._pixmap_cache:
            pm = QPixmap(img_path(filename))
            self._pixmap_cache[filename] = self._scale_pixmap(pm) if not pm.isNull() else pm
        return self._pixmap_cache[filename]

    def _get_movie(self, filename):
        if filename not in self._movie_cache:
            movie = QMovie(img_path(filename))
            if movie.isValid():
                movie.jumpToFrame(0)
                orig = movie.currentPixmap().size()
                if self.image_label_height > 0 and orig.height() > 0:
                    scaled_w = int(orig.width() * self.image_label_height / orig.height())
                    movie.setScaledSize(QSize(scaled_w, self.image_label_height))
                self._movie_cache[filename] = movie
            else:
                return None
        return self._movie_cache[filename]

    def _get_gif_label(self, index):
        while len(self._gif_labels) <= index:
            lbl = QLabel(self)
            lbl.setStyleSheet("background: transparent;")
            lbl.setAttribute(Qt.WA_TransparentForMouseEvents, True)
            lbl.hide(); self._gif_labels.append(lbl)
        return self._gif_labels[index]

    def _sync_gif_labels(self, gif_segments):
        for i, seg in enumerate(gif_segments):
            movie = self._get_movie(seg['content'])
            if movie is None: continue
            lbl = self._get_gif_label(i)
            if lbl.movie() is not movie:
                if lbl.movie(): lbl.movie().stop()
                lbl.setMovie(movie); movie.start()
            sz = movie.currentPixmap().size()
            if sz.isEmpty():
                sz = QSize(64, self.image_label_height if self.image_label_height > 0 else 64)
            lbl.resize(sz); lbl.move(seg['x'], seg['y']); lbl.show()
        for j in range(len(gif_segments), len(self._gif_labels)):
            lbl = self._gif_labels[j]
            if lbl.movie(): lbl.movie().stop()
            lbl.hide()

    def _hide_all_gifs(self):
        for lbl in self._gif_labels:
            if lbl.movie(): lbl.movie().stop()
            lbl.hide()

    def _build_label_segments(self, label_str):
        segments = []
        for line in label_str.split('\n'):
            stripped = line.strip()
            if stripped and is_image_label(stripped):
                ext  = os.path.splitext(stripped)[1].lower()
                kind = 'gif' if ext == '.gif' else 'image'
                segments.append({'kind': kind, 'content': stripped})
            else:
                segments.append({'kind': 'text', 'content': line})
        return segments

    def _compute_segment_layout(self, label_str, x1, start_y, w1):
        label_font    = QFont('Meiryo', self.label_font_size, QFont.Bold)
        label_metrics = QFontMetrics(label_font)
        line_h = label_metrics.lineSpacing()
        result = []; current_y = start_y
        for seg in self._build_label_segments(label_str):
            kind, content = seg['kind'], seg['content']
            if kind == 'image':
                pm = self._get_pixmap(content)
                if pm and not pm.isNull():
                    h = pm.height()
                    result.append({'kind': 'image', 'content': content, 'pixmap': pm,
                                   'x': x1 + (w1 - pm.width()) // 2, 'y': current_y, 'height': h})
                    current_y += h + 2
                else:
                    current_y += line_h
            elif kind == 'gif':
                movie = self._get_movie(content)
                if movie:
                    sz = movie.currentPixmap().size()
                    h  = sz.height() if sz.height() > 0 else (self.image_label_height or 32)
                    w  = sz.width()  if sz.width()  > 0 else h
                    result.append({'kind': 'gif', 'content': content,
                                   'x': x1 + (w1 - w) // 2, 'y': current_y, 'height': h})
                    current_y += h + 2
                else:
                    current_y += line_h
            else:
                text = content
                if text.strip():
                    tw = label_metrics.horizontalAdvance(text)
                    result.append({'kind': 'text', 'content': text,
                                   'x': x1 + (w1 - tw) // 2,
                                   'y': current_y + label_metrics.ascent(), 'height': line_h})
                current_y += line_h
        return result

    def update_display_text(self, time_str, label_str=""):
        self._current_label_str = label_str
        self.label.setText(f"{time_str}\n{label_str}" if label_str else f"{time_str}\n ")
        if label_str:
            time_metrics = QFontMetrics(QFont('Meiryo', self.timer_font_size, QFont.Bold))
            w1      = time_metrics.horizontalAdvance(time_str)
            start_y = time_metrics.ascent() + 10 + time_metrics.descent() - 3
            self._cached_segments = self._compute_segment_layout(label_str, 50, start_y, w1)
        else:
            self._cached_segments = []
        gif_segs = [s for s in self._cached_segments if s['kind'] == 'gif']
        if gif_segs: self._sync_gif_labels(gif_segs)
        else:        self._hide_all_gifs()
        self.update()

    def setupEvents(self):
        self._kb_signaler = TimerSignaler()
        self._kb_signaler.do_start_stop.connect(self.start_countdown)
        self._kb_signaler.do_log_a.connect(self.log_time_a)
        self._kb_signaler.do_log_ex.connect(self.log_time_ex)
        self._kb_signaler.do_log_loop.connect(self.log_time_b)
        self._kb_signaler.do_log_loopB.connect(self.log_time_loopB)
        self._kb_signaler.do_increment.connect(self.increment_counters)
        self._kb_signaler.do_decrement.connect(self.decrement_counters)
        self._kb_signaler.do_ai_count_2.connect(lambda: self.cycle_ai_count(2))
        self._kb_signaler.do_ai_count_3.connect(lambda: self.cycle_ai_count(3))
        self._loop_stop_timer = QTimer(self)
        self._loop_stop_timer.setSingleShot(True)
        self._loop_stop_timer.timeout.connect(self.stop_loop_only)
        self._register_key_hooks()

    def _register_key_hooks(self):
        for h in self._key_hooks:
            try: keyboard.unhook(h)
            except Exception: pass
        self._key_hooks = []

        bindings = [
            (self._kb_signaler.do_start_stop.emit, self.start_key),
            (self._kb_signaler.do_log_a.emit,      self.log_key),
            (self._kb_signaler.do_log_loop.emit,   self.log_key_loop),
            (self._kb_signaler.do_increment.emit,  self.increment_key),
            (self._kb_signaler.do_decrement.emit,  self.decrement_key),
        ]
        if self.log_key_ex:
            bindings.append((self._kb_signaler.do_log_ex.emit, self.log_key_ex))
        if self.log_key_loopB:
            bindings.append((self._kb_signaler.do_log_loopB.emit, self.log_key_loopB))
        if self.ai_count_key_2:
            bindings.append((self._kb_signaler.do_ai_count_2.emit, self.ai_count_key_2))
        if self.ai_count_key_3:
            bindings.append((self._kb_signaler.do_ai_count_3.emit, self.ai_count_key_3))

        parsed = []
        for emit, key_str in bindings:
            parts   = key_str.split('+')
            trigger = HotkeyCaptureThread._norm(parts[-1].strip().lower())
            mods    = [HotkeyCaptureThread._norm(m.strip().lower()) for m in parts[:-1]]
            parsed.append((trigger, mods, emit))

        def _on_event(e, _parsed=parsed):
            if e.event_type != keyboard.KEY_DOWN: return
            if not e.name: return
            name = HotkeyCaptureThread._norm(e.name.lower())
            for trigger, mods, emit in _parsed:
                if name == trigger and all(keyboard.is_pressed(m) for m in mods):
                    emit()

        self._key_hooks.append(keyboard.hook(_on_event))

    def set_hotkey(self, attr, value):
        setattr(self, attr, value)
        safe_set_config('config.ini', attr, value)
        self._register_key_hooks()

    def save_position_to(self, n, ox, oy):
        new_wx = ox + self.x()
        new_wy = oy + self.y()
        key_x = 'timer_x' if n == 1 else 'timer_x_2'
        key_y = 'timer_y' if n == 1 else 'timer_y_2'
        safe_set_config('config.ini', key_x, str(new_wx))
        safe_set_config('config.ini', key_y, str(new_wy))
        self.timer_x_list[n] = new_wx
        self.timer_y_list[n] = new_wy

    def save_position(self):
        if self.table_win:
            self.table_win.show_save_position_dialog()

    def open_color_dialog(self, attr):
        current = QColor(getattr(self, attr, '#ffffff'))
        color = QColorDialog.getColor(current, None, "カラーを選択", QColorDialog.ShowAlphaChannel)
        if not color.isValid(): return
        hex_color = color.name()
        setattr(self, attr, hex_color)
        safe_set_config('config.ini', attr, hex_color)
        self.update()

    def set_font_size(self, size):
        self.timer_font_size = size
        safe_set_config('config.ini', 'timer_font_size', str(size))
        self.update_display_text(self.label.text().split('\n')[0], self._current_label_str)

    def set_label_font_size(self, size):
        self.label_font_size = size
        safe_set_config('config.ini', 'label_font_size', str(size))
        self.update_display_text(self.label.text().split('\n')[0], self._current_label_str)

    def set_sound_volume(self, volume):
        self.sound_volume = volume
        safe_set_config('config.ini', 'sound_volume', str(volume))

    # ---- AIカウント ----
    def cycle_ai_count(self, max_val):
        self.ai_count_value = (self.ai_count_value % max_val) + 1
        self.update()

    def set_ai_count_visible(self, visible):
        self.ai_count_visible = visible
        safe_set_config('config.ini', 'ai_count_visible', '1' if visible else '0')
        self.update()

    def reload_from_config(self):
        config = configparser.ConfigParser()
        config.read('config.ini', encoding='utf-8')
        self.timer_font_color               = config.get('Settings', 'timer_font_color',               fallback=self.timer_font_color)
        self.timer_font_color_while_running = config.get('Settings', 'timer_font_color_while_running', fallback=self.timer_font_color_while_running)
        self.timer_font_color_ex            = config.get('Settings', 'timer_font_color_ex',            fallback=self.timer_font_color_ex)
        self.outline_color                  = config.get('Settings', 'outline_color',                  fallback=self.outline_color)
        self.timer_font_size                = config.getint('Settings', 'timer_font_size',              fallback=self.timer_font_size)
        self.label_font_size                = config.getint('Settings', 'label_font_size',              fallback=self.label_font_size)
        self.sound_volume                   = config.getint('Settings', 'sound_volume',                 fallback=self.sound_volume)
        self.start_key                      = config.get('Settings', 'start_key',                      fallback=self.start_key)
        self.log_key                        = config.get('Settings', 'log_key',                        fallback=self.log_key)
        self.log_key_ex                     = config.get('Settings', 'log_key_ex',                     fallback=self.log_key_ex)
        self.log_key_loop                   = config.get('Settings', 'log_key_loop',                   fallback=self.log_key_loop)
        self.log_key_loopB                  = config.get('Settings', 'log_key_loopB',                  fallback=self.log_key_loopB)
        self.increment_key                  = config.get('Settings', 'increment_key',                  fallback=self.increment_key)
        self.decrement_key                  = config.get('Settings', 'decrement_key',                  fallback=self.decrement_key)
        self.ai_count_key_2                 = config.get('Settings', 'ai_count_key_2',                 fallback=self.ai_count_key_2)
        self.ai_count_key_3                 = config.get('Settings', 'ai_count_key_3',                 fallback=self.ai_count_key_3)
        self.ai_count_value                 = 1
        self.ai_count_visible               = config.getboolean('Settings', 'ai_count_visible',    fallback=self.ai_count_visible)
        self.ai_count_font_color            = config.get('Settings',        'ai_count_font_color', fallback=self.ai_count_font_color)
        self.timer_x_list = [0,
            config.getint('Settings', 'timer_x',   fallback=self.timer_x_list[1]),
            config.getint('Settings', 'timer_x_2', fallback=self.timer_x_list[2])]
        self.timer_y_list = [0,
            config.getint('Settings', 'timer_y',   fallback=self.timer_y_list[1]),
            config.getint('Settings', 'timer_y_2', fallback=self.timer_y_list[2])]
        self._register_key_hooks()
        self.update_display_text(self.label.text().split('\n')[0], self._current_label_str)

    def setupMediaPlayer(self):
        def _p(f):
            p = QMediaPlayer()
            p.setMedia(QMediaContent(QUrl.fromLocalFile(sound_path(f)))); return p
        self.start_player    = _p('start.mp3')
        self.bomb_player     = _p('activation.mp3')
        self.record_player_a = _p('record_a.mp3')
        self.record_player_b = _p('record_b.mp3')

    def _start_hotkey_capture(self, action_name, attr, target):
        if self._capture_overlay is not None:
            try: self._capture_overlay.cancel()
            except Exception: pass
        self._capture_overlay = HotkeyCaptureOverlay(
            action_name,
            lambda key, a=attr, t=target: t.set_hotkey(a, key),
            parent=None)
        self._capture_overlay.show()

    def _build_color_menu(self, parent_menu):
        color_menu = parent_menu.addMenu("フォントカラーの設定")
        for label, attr in [("通常時",       "timer_font_color"),
                             ("動作中",       "timer_font_color_while_running"),
                             ("EX",          "timer_font_color_ex"),
                             ("AIカウント",   "ai_count_font_color"),
                             ("縁取り",       "outline_color")]:
            cur = getattr(self, attr, '#ffffff')
            color_menu.addAction(_make_color_icon(cur), f"{label}  {cur}").triggered.connect(
                lambda checked, a=attr: self.open_color_dialog(a))

    def _build_settings_save_menu(self, menu):
        twr = self.table_win; tmr = self
        sp  = menu.addMenu("設定・保存")

        bak_path   = 'config.ini.bak'
        bak_exists = os.path.isfile(bak_path)
        act_restore = sp.addAction("設定を戻す")
        act_restore.triggered.connect(
            twr.show_restore_config_dialog if twr else lambda: None)
        act_restore.setEnabled(bak_exists)
        sp.addSeparator()

        sp.addAction("現在の位置を保存").triggered.connect(
            twr.show_save_position_dialog if twr else lambda: None)

        if self.main_win:
            cur_op = self.main_win._opacity
            op_menu = sp.addMenu("透過率の設定")
            for v in range(100, 0, -10):
                lbl = f"{'▶ ' if v == cur_op else '　'}{v}%"
                op_menu.addAction(lbl).triggered.connect(
                    lambda checked, val=v: self.main_win.set_opacity(val))

        if twr:
            xlsx_menu = sp.addMenu("テーブルの表示設定")
            for align, label in (('right', '右揃え'), ('left', '左揃え')):
                lbl = f"{'▶ ' if twr.table_align == align else '　'}{label}"
                xlsx_menu.addAction(lbl).triggered.connect(
                    lambda checked, a=align: twr.set_table_align(a))
            xlsx_menu.addSeparator()
            vis = getattr(twr, 'table_visible', True)
            xlsx_menu.addAction(f"{'▶ ' if vis     else '　'}表示").triggered.connect(
                lambda: twr.set_table_visible(True))
            xlsx_menu.addAction(f"{'▶ ' if not vis else '　'}非表示").triggered.connect(
                lambda: twr.set_table_visible(False))

        font_menu = sp.addMenu("フォントサイズの設定")
        timer_font_menu = font_menu.addMenu("タイマー")
        for fs in range(8, 81, 4):
            lbl = f"{'▶ ' if fs == tmr.timer_font_size else '　'}{fs}pt"
            timer_font_menu.addAction(lbl).triggered.connect(
                lambda checked, v=fs: tmr.set_font_size(v))
        label_font_menu = font_menu.addMenu("ラベル")
        for fs in range(6, 33, 2):
            lbl = f"{'▶ ' if fs == tmr.label_font_size else '　'}{fs}pt"
            label_font_menu.addAction(lbl).triggered.connect(
                lambda checked, v=fs: tmr.set_label_font_size(v))

        vol_menu = sp.addMenu("音量の設定")
        for v in range(100, -1, -10):
            lbl = f"{'▶ ' if v == tmr.sound_volume else '　'}{v}%"
            vol_menu.addAction(lbl).triggered.connect(
                lambda checked, val=v: tmr.set_sound_volume(val))

        self._build_color_menu(sp)

        hk_menu = sp.addMenu("Hotkeyの設定")
        hotkey_items = [
            ("タイマー開始/停止", "start_key",         tmr),
            ("タイマー記録",            "log_key",            tmr),
            ("タイマー記録[EX表示]",           "log_key_ex",         tmr),
            ("タイマー記録[ループA開始]",       "log_key_loop",       tmr),
            ("タイマー記録[ループB開始]",       "log_key_loopB",      tmr),
            ("タイマー ズレ修正[増加]",    "increment_key",      tmr),
            ("タイマー ズレ修正[減少]",    "decrement_key",      tmr),
            ("シート切替",        "switch_key",         twr),
            ("プリセット切替",    "preset_switch_key",  twr),
            ("位置リセット",      "reset_position_key", twr),
        ]
        for name, attr, target in hotkey_items:
            if target is None: continue
            cur = getattr(target, attr, "") or "未設定"
            hk_menu.addAction(f"{name}  [{cur}]").triggered.connect(
                lambda checked, n=name, a=attr, t=target: self._start_hotkey_capture(n, a, t))
        hk_menu.addSeparator()
        for ai_name, ai_attr in [("AIカウント (1-2)",   "ai_count_key_2"),
                                  ("AIカウント (1-2-3)", "ai_count_key_3")]:
            ai_cur = getattr(tmr, ai_attr, "") or "未設定"
            hk_menu.addAction(f"{ai_name}  [{ai_cur}]").triggered.connect(
                lambda checked, n=ai_name, a=ai_attr, t=tmr: self._start_hotkey_capture(n, a, t))

        # ---- AIカウント ----
        ai_menu = sp.addMenu("AIカウントの表示設定")
        vis_ai = tmr.ai_count_visible
        ai_menu.addAction(f"{'▶ ' if vis_ai     else '　'}表示").triggered.connect(
            lambda: tmr.set_ai_count_visible(True))
        ai_menu.addAction(f"{'▶ ' if not vis_ai else '　'}非表示").triggered.connect(
            lambda: tmr.set_ai_count_visible(False))

    def showContextMenu(self, pos):
        menu = QMenu(self)
        menu.setStyleSheet("QMenu { color: white; background-color: rgba(50, 50, 50, 150); border: 1px solid #555; }"
                           "QMenu::item:selected { background-color: rgba(100, 100, 100, 200); }")
        if self.table_win:
            self.table_win._add_common_top_items(menu)
        self._build_settings_save_menu(menu)
        menu.addSeparator()
        menu.addAction("閉じる").triggered.connect(QApplication.instance().quit)
        menu.addAction("キャンセル")
        menu.exec_(self.mapToGlobal(pos))

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self._drag_start = event.globalPos(); event.accept()

    def mouseMoveEvent(self, event):
        if event.buttons() == Qt.LeftButton and self._drag_start is not None:
            delta = event.globalPos() - self._drag_start
            self._drag_start = event.globalPos()
            self.move(self.pos() + delta); event.accept()

    def mouseReleaseEvent(self, event):
        if event.button() == Qt.LeftButton:
            self._drag_start = None

    def time_to_seconds(self, time_str):
        m, s = map(int, time_str.split(':'))
        return m * 60 + s

    def seconds_to_time(self, seconds):
        m, s = divmod(int(max(0, seconds)), 60)
        return f'{m:02d}:{s:02d}'

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        painter.setRenderHint(QPainter.TextAntialiasing)
        time_font    = QFont('Meiryo', self.timer_font_size, QFont.Bold)
        label_font   = QFont('Meiryo', self.label_font_size, QFont.Bold)
        lines        = self.label.text().split('\n')
        time_metrics = QFontMetrics(time_font)
        is_running   = self.timer_thread and self.timer_thread.isRunning()
        timer_color  = QColor(self.timer_font_color_while_running) if is_running else QColor(self.timer_font_color)
        line1 = lines[0]
        w1    = time_metrics.horizontalAdvance(line1)
        x1, y1 = 50, time_metrics.ascent() + 10

        timer_path = QPainterPath()
        timer_path.addText(x1, y1, time_font, line1)
        label_text_path = QPainterPath()
        for seg in self._cached_segments:
            if seg['kind'] == 'text':
                label_text_path.addText(seg['x'], seg['y'], label_font, seg['content'])
            elif seg['kind'] == 'image':
                painter.drawPixmap(seg['x'], seg['y'], seg['pixmap'])

        full_path = QPainterPath(timer_path)
        full_path.addPath(label_text_path)
        pen = QPen(QColor(self.outline_color))
        pen.setWidthF(5.0); pen.setJoinStyle(Qt.RoundJoin)
        painter.setPen(pen); painter.drawPath(full_path)
        painter.fillPath(timer_path,      QBrush(timer_color))
        painter.fillPath(label_text_path, QBrush(timer_color))

        if self.ex_log is not None:
            ex_time_str = self.seconds_to_time(self.ex_log[0])
            ex_x_base   = x1 + w1 + 16
            ex_path = QPainterPath()
            ex_path.addText(ex_x_base, y1, time_font, ex_time_str)
            for seg in self._cached_ex_segments:
                if seg['kind'] == 'text':
                    ex_path.addText(seg['x'], seg['y'], label_font, seg['content'])
                elif seg['kind'] == 'image':
                    painter.drawPixmap(seg['x'], seg['y'], seg['pixmap'])
            ex_pen = QPen(QColor(self.outline_color))
            ex_pen.setWidthF(5.0); ex_pen.setJoinStyle(Qt.RoundJoin)
            painter.setPen(ex_pen); painter.drawPath(ex_path)
            painter.fillPath(ex_path, QBrush(QColor(self.timer_font_color_ex)))

        # ---- AIカウント描画 ----
        if self.ai_count_visible:
            ai_font_size = max(10, self.timer_font_size // 3)
            ai_font      = QFont('Meiryo', ai_font_size, QFont.Bold)
            ai_metrics   = QFontMetrics(ai_font)
            ai_spacing   = ai_metrics.horizontalAdvance("8") + 6
            ai_y         = ai_metrics.ascent() - ai_font_size // 3
            max_count    = 3
            cur          = self.ai_count_value
            ai_outline   = QPen(QColor(self.outline_color))
            ai_outline.setWidthF(3.0); ai_outline.setJoinStyle(Qt.RoundJoin)
            painter.setPen(ai_outline)
            for i in range(1, max_count + 1):
                ai_x    = x1 + (i - 1) * ai_spacing
                ai_path = QPainterPath()
                ai_path.addText(ai_x, ai_y, ai_font, str(i))
                if i == cur:
                    painter.drawPath(ai_path)
                    painter.fillPath(ai_path, QBrush(QColor(self.ai_count_font_color)))

        painter.end()

    def update_time(self, current_time):
        if self.ex_log is not None and self.activation_ex > 0 and current_time <= self.ex_log[0]:
            self.ex_log = None; self._cached_ex_segments = []
        while self.combined_logs:
            target_time, target_label = self.combined_logs[0]
            if self.play_bomb_sound and self.last_played_target != target_time:
                if current_time <= (target_time + self.sound_delay):
                    self.bomb_player.stop(); self.bomb_player.setVolume(self.sound_volume); self.bomb_player.play()
                    self.last_played_target = target_time
            if current_time > target_time:
                self.update_display_text(self.seconds_to_time(target_time), target_label); break
            else:
                self.combined_logs.pop(0)
        else:
            if not (self.timer_thread and self.timer_thread.isRunning()):
                self.update_display_text(self.initial_time, "")
        self.update()

    def setupLogs(self):
        self.combined_logs      = []
        self.log_time_b_counter = 0
        self.ex_log             = None

    def clear_logs_only(self):
        self.current_loopA_index = 0; self.current_loopB_index = 0
        self.combined_logs.clear()
        self.log_time_b_counter = 0; self.last_played_target = -1
        self.ex_log = None; self._cached_ex_segments = []

    def start_countdown(self):
        if self.timer_thread is not None and self.timer_thread.isRunning():
            self.reset_timer_and_logs()
        else:
            start_secs = self.time_to_seconds(self.initial_time)
            self.clear_logs_only()
            self.timer_thread = TimerThread(start_secs, self.reset_timer_and_logs)
            self.timer_thread.timer_updated.connect(self.update_time)
            self.timer_thread.start(); self.update()
            if self.play_start_sound:
                self.start_player.setVolume(self.sound_volume); self.start_player.play()

    def log_time_a(self):
        if self.timer_thread and self.timer_thread.isRunning():
            recorded_time = self.timer_thread.current_time - self.activation_normal
            self.combined_logs.append((recorded_time, self.log_label))
            self.combined_logs.sort(key=lambda x: x[0], reverse=True)
            if self.play_record_sound:
                self.record_player_a.setVolume(self.sound_volume); self.record_player_a.play()

    def log_time_b(self):
        if self.timer_thread and self.timer_thread.isRunning():
            self.log_time_b_sc()
            if self.play_record_sound:
                self.record_player_b.setVolume(self.sound_volume); self.record_player_b.play()
            if self.loopA_stop_time > 0:
                self._schedule_loop_stop(self.loopA_stop_time)

    def log_time_loopB(self):
        if self.timer_thread and self.timer_thread.isRunning():
            self.log_time_b_br()
            if self.play_record_sound:
                self.record_player_b.setVolume(self.sound_volume); self.record_player_b.play()
            if self.loopB_stop_time > 0:
                self._schedule_loop_stop(self.loopB_stop_time)

    def _schedule_loop_stop(self, delay_secs):
        self._loop_stop_timer.stop()
        self._loop_stop_timer.start(int(delay_secs * 1000))

    def log_time_ex(self):
        if self.timer_thread and self.timer_thread.isRunning():
            recorded_time = self.timer_thread.current_time - self.activation_ex
            self.ex_log = (recorded_time, self.log_label_ex)
            self._rebuild_ex_cache()
            if self.play_record_sound:
                self.record_player_a.setVolume(self.sound_volume); self.record_player_a.play()
            self.update()

    def _rebuild_ex_cache(self):
        if self.ex_log is None:
            self._cached_ex_segments = []; return
        ex_time_str  = self.seconds_to_time(self.ex_log[0])
        ex_label_str = (self.ex_log[1] or "")
        time_metrics = QFontMetrics(QFont('Meiryo', self.timer_font_size, QFont.Bold))
        x1        = 50
        w1        = time_metrics.horizontalAdvance(self.label.text().split('\n')[0])
        ex_x_base = x1 + w1 + 16
        ex_w1     = time_metrics.horizontalAdvance(ex_time_str)
        y1        = time_metrics.ascent() + 10
        ex_current_y = y1 + time_metrics.descent() - 3
        if ex_label_str:
            self._cached_ex_segments = self._compute_segment_layout(
                ex_label_str, ex_x_base, ex_current_y, ex_w1)
        else:
            self._cached_ex_segments = []

    def log_time_b_sc(self):
        recorded_time = self.timer_thread.current_time - self.activation_loopA
        current_loopA_time = recorded_time; idx = self.current_loopA_index
        while current_loopA_time > 0:
            label = self.loopA_labels[idx % len(self.loopA_labels)]
            self.combined_logs.append((current_loopA_time, label))
            current_loopA_time -= self.loopA_times[idx % len(self.loopA_times)]; idx += 1
        self.current_loopA_index = (self.current_loopA_index + 1) % len(self.loopA_times)
        self.combined_logs.sort(key=lambda x: x[0], reverse=True)

    def log_time_b_br(self):
        recorded_time = self.timer_thread.current_time - self.activation_loopB
        current_loopB_time = recorded_time; idx = self.current_loopB_index; loopB_entries = []
        while current_loopB_time > 0:
            label = self.loopB_labels[idx % len(self.loopB_labels)]
            loopB_entries.append((current_loopB_time, label))
            current_loopB_time -= self.loopB_times[idx % len(self.loopB_times)]; idx += 1
        self.current_loopB_index = (self.current_loopB_index + 1) % len(self.loopB_times)
        self.combined_logs = [x for x in self.combined_logs if x[0] > recorded_time]
        self.combined_logs.extend(loopB_entries)
        self.combined_logs.sort(key=lambda x: x[0], reverse=True)

    def increment_counters(self):
        if self.timer_thread and self.timer_thread.isRunning():
            self.combined_logs = [(t + 1, l) for t, l in self.combined_logs]
            self.update_time(self.timer_thread.current_time)
        else:
            new_secs = self.time_to_seconds(self.initial_time) + 10
            self.initial_time = self.seconds_to_time(new_secs)
            self.update_display_text(self.initial_time, "")
        self.update()

    def decrement_counters(self):
        if self.timer_thread and self.timer_thread.isRunning():
            self.combined_logs = [(max(0, t - 1), l) for t, l in self.combined_logs]
            self.update_time(self.timer_thread.current_time)
        else:
            new_secs = max(0, self.time_to_seconds(self.initial_time) - 10)
            self.initial_time = self.seconds_to_time(new_secs)
            self.update_display_text(self.initial_time, "")
        self.update()

    def stop_loop_only(self):
        self._loop_stop_timer.stop()
        self.clear_logs_only()
        self.update()

    def reset_timer_and_logs(self):
        self._loop_stop_timer.stop()
        if self.timer_thread is not None:
            self.timer_thread.stop(); self.timer_thread.wait(); self.timer_thread = None
        self.initial_time = self.reset_time
        self.clear_logs_only()
        self.update_display_text(self.initial_time, "")
        self.update()


# ============================================================
# メインウィンドウ
# ============================================================
class MainWindow(QWidget):
    CONFIG_PATH = 'config.ini'

    def __init__(self):
        super().__init__()
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint | Qt.Window)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setAutoFillBackground(False)
        desktop = QApplication.desktop()
        vg = desktop.screenGeometry(0)
        for i in range(1, desktop.screenCount()):
            vg = vg.united(desktop.screenGeometry(i))
        self.setGeometry(vg)
        config = configparser.ConfigParser()
        config.read(self.CONFIG_PATH, encoding='utf-8')
        self._opacity = max(10, min(100, config.getint('Settings', 'opacity', fallback=100)))
        self.setWindowOpacity(self._opacity / 100)

    def set_opacity(self, value):
        self._opacity = max(10, min(100, value))
        self.setWindowOpacity(self._opacity / 100)
        safe_set_config(self.CONFIG_PATH, 'opacity', str(self._opacity))

    def paintEvent(self, event): pass


# ============================================================
# アプリケーションクラス
# ============================================================
class XTimerApp:
    def __init__(self):
        presets = load_presets('config.ini')
        self.main_win  = MainWindow()
        self.timer_win = TransparentWindow(parent=self.main_win)
        self.table_win = ExcelTableWindow('config.ini', presets, parent=self.main_win)
        self.table_win.set_timer_window(self.timer_win)
        self.timer_win.set_table_window(self.table_win)
        self.timer_win.set_main_window(self.main_win)
        self.table_win.set_main_window(self.main_win)
        self.timer_win._apply_initial_position()

    def show(self):
        self.main_win.show()
        self.timer_win.show()
        self.table_win.show()


# ============================================================
# エントリーポイント
# ============================================================
if __name__ == '__main__':
    def handle_exception(exc_type, exc_value, exc_traceback):
        import traceback
        print("".join(traceback.format_exception(exc_type, exc_value, exc_traceback)))
    sys.excepthook = handle_exception

    try:
        from PyQt5.QtCore import qInstallMessageHandler
        def qt_message_handler(mode, context, message):
            print(f"[Qt] {message}")
        qInstallMessageHandler(qt_message_handler)
    except Exception: pass

    if sys.platform == 'win32':
        myappid = 'dqx.Xtimer.v1.0'
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)

    app = QApplication(sys.argv)
    ico = img_path("XTimer128.ico")
    if getattr(sys, 'frozen', False):
        ico = os.path.join(sys._MEIPASS, ico)
    app.setWindowIcon(QIcon(ico))

    X_app = XTimerApp()
    X_app.show()
    sys.exit(app.exec_())